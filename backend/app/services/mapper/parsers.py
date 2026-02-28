"""Dataset readers for Smart Mapper."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None


def normalize_header(value: str) -> str:
    normalized = re.sub(r"[^0-9a-zA-Z]+", "_", value.strip().lower())
    return re.sub(r"_+", "_", normalized).strip("_")


class DatasetReader:
    def __init__(
        self,
        file_path: Path,
        file_type: str,
        sheet: str | None = None,
        header_row: int = 1,
    ) -> None:
        self.file_path = file_path
        self.file_type = file_type
        self.sheet = sheet
        self.header_row = header_row

    def list_sheets(self) -> list[tuple[str, int | None, int | None]]:
        if self.file_type == "csv":
            return [(self.file_path.name, None, None)]
        if openpyxl is None:
            return []
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append((name, ws.max_row, ws.max_column))
        wb.close()
        return sheets

    def iter_rows(self) -> Iterable[list[object | None]]:
        if self.file_type == "csv":
            return self._iter_csv()
        return self._iter_xlsx()

    def read_rows(self, max_rows: int | None = None) -> list[list[object | None]]:
        rows = []
        for idx, row in enumerate(self.iter_rows(), start=1):
            rows.append(row)
            if max_rows is not None and idx >= max_rows:
                break
        return rows

    def _iter_csv(self) -> Iterable[list[object | None]]:
        with self.file_path.open("rb") as raw:
            sample = raw.read(8192)
        try:
            dialect = csv.Sniffer().sniff(sample.decode("utf-8", errors="ignore"))
        except csv.Error:
            dialect = csv.get_dialect("excel")

        with self.file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, dialect)
            for row in reader:
                yield [cell if cell != "" else None for cell in row]

    def _iter_xlsx(self) -> Iterable[list[object | None]]:
        if openpyxl is None:
            raise RuntimeError("openpyxl is required to parse .xlsx files")
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        sheet_name = self.sheet or wb.sheetnames[0]
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            yield [cell for cell in row]
        wb.close()


def infer_value_type(value: object) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (datetime, date)):
        return "date"
    text = str(value).strip()
    if not text:
        return "empty"
    lowered = text.lower()
    if lowered in {"true", "false", "yes", "no", "0", "1"}:
        return "boolean"
    try:
        float(text.replace(",", ""))
        return "number"
    except ValueError:
        pass
    try:
        datetime.fromisoformat(text)
        return "date"
    except ValueError:
        return "string"


def summarize_types(values: list[object | None]) -> str:
    types = {infer_value_type(v) for v in values if infer_value_type(v) != "empty"}
    if not types:
        return "empty"
    if len(types) == 1:
        return types.pop()
    return "mixed"


def build_examples(values: list[object | None], limit: int = 5) -> list[str]:
    seen: set[str] = set()
    examples: list[str] = []
    for value in values:
        if value is None:
            continue
        text = str(value)
        if text in seen:
            continue
        seen.add(text)
        examples.append(text)
        if len(examples) >= limit:
            break
    return examples
